import openpyxl
from Materials import Materials
import sys

직업 = ['목수', '대장', '갑주', '보석', '가죽', '재봉', '연금', '요리']

wb = openpyxl.load_workbook('./docs/제작수첩DB.xlsx')

#TotMat = []
mdb = Materials()

# PRE-SCAN items
for sheet in wb:
    #print("### ", sheet.title, sheet.max_row, sheet.max_column)

    #mdb = Materials(sheet.max_row, sheet.title)
    mdb.insert(sheet.max_row, sheet.title)
    materials = (sheet.max_row, sheet.title)

    quantity = [0, 0, 0, 0, 0, 0]
    item = ["", "", "", "", "", ""]

    mr = sheet.max_row
    mc = sheet.max_column

    for r in range(2, mr):
        if(sheet.cell(r, 1).value == None):
            continue

        for i in range(0, 5):
            cq = i*2+5
            ci = i*2+6
            #quantity[i] = sheet.cell(r,c1).value
            #item[i] = sheet.cell(r,c2).value
            Q = sheet.cell(r, cq).value
            NAME = sheet.cell(r, ci).value

            if(Q != None):
                quantity[i] = Q

                if(NAME != None):
                    item[i] = NAME

                #print(sheet.title,",",r,",",quantity[i],",",item[i])
                mdb.insert(quantity[i], item[i])
                #print(quantity[i], item[i])
            else:
                break

mdb.Print()

# OUTPUT #
##create new workbook
#oDB = openpyxl.Workbook()

# copying the cell values from source
# excel file to destination excel file

#oDB.save('extraction.xlsx')
