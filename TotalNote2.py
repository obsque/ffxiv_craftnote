from ctypes import sizeof
import openpyxl
import pandas as pd
from item import ITEM
import sys

FILTER_LEVEL = 55

직업 = ['목수', '대장', '갑주', '보석', '가죽', '재봉', '연금', '요리']

wb = openpyxl.load_workbook('./docs/제작수첩DB.xlsx')

#TotMat = []
product_dict = {}


# PRE-SCAN items - Dictionary
for sheet in wb:

    mr = sheet.max_row
    mc = sheet.max_column

    quantity = [0, 0, 0, 0, 0, 0]
    item = ["", "", "", "", "", ""]

    # ADD to dict
    for r in range(2, mr):
        product = sheet.cell(r, 4).value
        # if name not in item_dict:
        product_dict[product] = {}

        for i in range(0, 5):
            cq = i*2+5
            ci = i*2+6

            Q = sheet.cell(r,cq).value
            MATERIAL = sheet.cell(r,ci).value
            
            if(Q != None):
                quantity[i] = Q

                if(MATERIAL != None):
                    item[i] = MATERIAL
                else:
                    MATERIAL = item[i]

                # ADD to dict
                product_dict[product][MATERIAL] = Q

# print(item_dict)

################################################################################

cart_dict = {}

def AddItem(key, quantity):
    if key in cart_dict:
        cart_dict[key] += quantity
    else:
        cart_dict[key] = quantity

def AddMaterials(key, quantity):
    # TBD Recursive
    if key in product_dict:       
        AddItem(key, quantity)
        for material in list(product_dict[key]):
            AddMaterials(material, product_dict[key][material]*quantity)
    else:
            AddItem(key, quantity)

# collect materials
for sheet in wb:

    quantity = [0, 0, 0, 0, 0, 0]
    item = ["", "", "", "", "", ""]

    mr = sheet.max_row
    mc = sheet.max_column

    for r in range(2, mr):

        if (sheet.cell(r, 3).value):
            reqlevel = sheet.cell(r, 3).value        

        # check : need to crafting
        if(sheet.cell(r, 1).value == None):
            continue
        
        if reqlevel > FILTER_LEVEL:
            continue

        # maybe never
        product = sheet.cell(r, 4).value

        # for materials
        for material in list(product_dict[product]):
            AddMaterials(material, product_dict[product][material])

# OUTPUT #
data = pd.DataFrame(list(cart_dict.items()), columns=['Name', 'Quantity'])
print(data)
output = data.loc[data['Quantity'] > 0]

output.to_excel("output.xlsx")
